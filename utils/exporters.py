"""
utils/exporters.py
===================
Generación de reportes exportables en PDF y Excel a partir de los
datos del sistema (Capítulo II: "Reportes automatizados exportables").

PDF: usa reportlab (Platypus) para reportes con encabezado institucional,
     tablas, indicadores y gráficos incrustados.
Excel: usa openpyxl para exportar tablas de datos con formato
     profesional, listas para análisis posterior por el usuario.
"""

import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import config
from controllers import estadisticas_controller as ec
from models.militar import PersonalMilitar

# ---------------------------------------------------------------------
# Colores institucionales (coherentes con config.py y graficos_controller.py)
# ---------------------------------------------------------------------
COLOR_PRIMARIO_HEX = "1B4332"
COLOR_SECUNDARIO_HEX = "40916C"
COLOR_ACENTO_HEX = "D4AF37"


# =======================================================================
# EXPORTACIÓN A EXCEL
# =======================================================================

def _estilo_encabezado(celda):
    celda.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    celda.fill = PatternFill("solid", start_color=COLOR_PRIMARIO_HEX)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = Border(*[Side(style="thin", color="999999")] * 4)


def _autoajustar_columnas(hoja, columnas_anchos: dict) -> None:
    for col_letra, ancho in columnas_anchos.items():
        hoja.column_dimensions[col_letra].width = ancho


def exportar_personal_excel(lista_personal: List[PersonalMilitar], ruta_salida: Optional[str] = None) -> str:
    """
    Exporta una lista de registros de personal a un archivo Excel con
    formato profesional (encabezados resaltados, columnas ajustadas,
    filtros automáticos). Retorna la ruta del archivo generado.
    """
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Personal Militar"

    encabezados = [
        "Cédula", "Nombres", "Apellidos", "Edad", "Género",
        "Nivel Educativo", "Grado", "Municipio", "Parroquia",
        "Estatus", "Fecha Registro", "Teléfono", "Dirección",
    ]
    hoja.append(encabezados)
    for col_idx in range(1, len(encabezados) + 1):
        _estilo_encabezado(hoja.cell(row=1, column=col_idx))

    fuente_normal = Font(name="Arial", size=10)
    for fila_idx, p in enumerate(lista_personal, start=2):
        valores = [
            p.cedula, p.nombres, p.apellidos, p.edad, p.genero,
            p.nivel_educativo or "", p.grado or "", p.municipio,
            p.parroquia or "", p.estatus, p.fecha_registro,
            p.telefono or "", p.direccion or "",
        ]
        hoja.append(valores)
        for col_idx in range(1, len(valores) + 1):
            celda = hoja.cell(row=fila_idx, column=col_idx)
            celda.font = fuente_normal
            if fila_idx % 2 == 0:
                celda.fill = PatternFill("solid", start_color="F2F2F2")

    anchos = {"A": 12, "B": 18, "C": 20, "D": 8, "E": 12, "F": 26,
              "G": 16, "H": 14, "I": 22, "J": 18, "K": 14, "L": 16, "M": 28}
    _autoajustar_columnas(hoja, anchos)
    hoja.freeze_panes = "A2"  # mantiene encabezados visibles al desplazar
    hoja.auto_filter.ref = hoja.dimensions  # filtros automáticos en encabezados

    if ruta_salida is None:
        nombre_archivo = f"personal_militar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_salida = os.path.join(config.REPORTS_DIR, nombre_archivo)

    wb.save(ruta_salida)
    return ruta_salida


def exportar_resumen_estadistico_excel(ruta_salida: Optional[str] = None) -> str:
    """
    Exporta un Excel con múltiples hojas: resumen general, distribución
    por municipio, por parroquia, por estatus, y tabla cruzada
    género x municipio. Ideal como anexo estadístico para el trabajo escrito.
    """
    wb = Workbook()
    wb.remove(wb.active)  # se crean hojas con nombres propios, no la "Sheet" por defecto

    # --- Hoja 1: Resumen general (KPIs) ---
    hoja_resumen = wb.create_sheet("Resumen General")
    resumen = ec.resumen_general()
    hoja_resumen.append(["Indicador", "Valor"])
    for col_idx in (1, 2):
        _estilo_encabezado(hoja_resumen.cell(row=1, column=col_idx))
    etiquetas = {
        "total_registros": "Total de registros",
        "total_masculino": "Total masculino",
        "total_femenino": "Total femenino",
        "edad_promedio": "Edad promedio",
        "edad_minima": "Edad mínima",
        "edad_maxima": "Edad máxima",
        "total_municipios_con_registros": "Municipios con registros",
    }
    for clave, etiqueta in etiquetas.items():
        hoja_resumen.append([etiqueta, resumen.get(clave, "")])
    hoja_resumen.column_dimensions["A"].width = 28
    hoja_resumen.column_dimensions["B"].width = 16

    # --- Hoja 2: Distribución por municipio ---
    hoja_mun = wb.create_sheet("Por Municipio")
    hoja_mun.append(["Municipio", "Total Registros"])
    for col_idx in (1, 2):
        _estilo_encabezado(hoja_mun.cell(row=1, column=col_idx))
    for municipio, total in ec.distribucion_por_municipio().items():
        hoja_mun.append([municipio, total])
    hoja_mun.column_dimensions["A"].width = 24
    hoja_mun.column_dimensions["B"].width = 16

    # --- Hoja 3: Distribución por parroquia ---
    hoja_par = wb.create_sheet("Por Parroquia")
    hoja_par.append(["Parroquia", "Total Registros"])
    for col_idx in (1, 2):
        _estilo_encabezado(hoja_par.cell(row=1, column=col_idx))
    for parroquia, total in ec.distribucion_por_parroquia().items():
        hoja_par.append([parroquia, total])
    hoja_par.column_dimensions["A"].width = 30
    hoja_par.column_dimensions["B"].width = 16

    # --- Hoja 4: Distribución por estatus ---
    hoja_est = wb.create_sheet("Por Estatus")
    hoja_est.append(["Estatus de Participación", "Total Registros"])
    for col_idx in (1, 2):
        _estilo_encabezado(hoja_est.cell(row=1, column=col_idx))
    for estatus, total in ec.distribucion_por_estatus().items():
        hoja_est.append([estatus, total])
    hoja_est.column_dimensions["A"].width = 28
    hoja_est.column_dimensions["B"].width = 16

    # --- Hoja 5: Tabla cruzada Género x Municipio ---
    hoja_cruzada = wb.create_sheet("Genero x Municipio")
    tabla_cruzada = ec.tabla_cruzada_genero_municipio()
    if not tabla_cruzada.empty:
        encabezado = ["Municipio"] + list(tabla_cruzada.columns)
        hoja_cruzada.append(encabezado)
        for col_idx in range(1, len(encabezado) + 1):
            _estilo_encabezado(hoja_cruzada.cell(row=1, column=col_idx))
        for municipio, fila in tabla_cruzada.iterrows():
            hoja_cruzada.append([municipio] + list(fila.values))
        hoja_cruzada.column_dimensions["A"].width = 24
        for idx in range(2, len(encabezado) + 1):
            hoja_cruzada.column_dimensions[get_column_letter(idx)].width = 14

    if ruta_salida is None:
        nombre_archivo = f"resumen_estadistico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_salida = os.path.join(config.REPORTS_DIR, nombre_archivo)

    wb.save(ruta_salida)
    return ruta_salida


# =======================================================================
# EXPORTACIÓN A PDF
# =======================================================================

def _estilos_pdf():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TituloInstitucional", parent=styles["Title"],
        fontSize=15, textColor=colors.HexColor("#" + COLOR_PRIMARIO_HEX),
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="Subtitulo", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#555555"),
        alignment=1, spaceAfter=12,
    ))
    styles.add(ParagraphStyle(
        name="SeccionTitulo", parent=styles["Heading2"],
        fontSize=12, textColor=colors.HexColor("#" + COLOR_PRIMARIO_HEX),
        spaceBefore=14, spaceAfter=6,
    ))
    return styles


def _encabezado_institucional(styles) -> list:
    """Construye los elementos de encabezado institucional, reutilizados en todos los PDF."""
    elementos = []
    elementos.append(Paragraph(config.NOMBRE_SISTEMA, styles["TituloInstitucional"]))
    elementos.append(Paragraph(config.SUBTITULO_SISTEMA, styles["Subtitulo"]))
    elementos.append(Paragraph(config.UNIDAD_MILITAR, styles["Subtitulo"]))
    elementos.append(Spacer(1, 0.3 * cm))
    fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    elementos.append(Paragraph(f"Fecha de generación: {fecha_str}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))
    return elementos


def _tabla_estilizada(data: list, anchos_col: Optional[list] = None) -> Table:
    tabla = Table(data, colWidths=anchos_col)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + COLOR_PRIMARIO_HEX)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tabla


def generar_reporte_estadistico_pdf(ruta_salida: Optional[str] = None, incluir_graficos: bool = True) -> str:
    """
    Genera un reporte PDF completo con: resumen general (KPIs),
    distribución por género, edad, nivel educativo, municipio,
    parroquia y estatus de participación. Si incluir_graficos=True,
    incrusta los gráficos generados con matplotlib.

    Retorna la ruta del archivo generado.
    """
    if ruta_salida is None:
        nombre_archivo = f"reporte_estadistico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta_salida = os.path.join(config.REPORTS_DIR, nombre_archivo)

    styles = _estilos_pdf()
    doc = SimpleDocTemplate(
        ruta_salida, pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
    )
    elementos = _encabezado_institucional(styles)

    # --- Resumen general (KPIs) ---
    elementos.append(Paragraph("Resumen General de Participación", styles["SeccionTitulo"]))
    resumen = ec.resumen_general()
    datos_resumen = [
        ["Indicador", "Valor"],
        ["Total de registros", str(resumen["total_registros"])],
        ["Total masculino", str(resumen["total_masculino"])],
        ["Total femenino", str(resumen["total_femenino"])],
        ["Edad promedio", f"{resumen['edad_promedio']} años"],
        ["Rango de edad", f"{resumen['edad_minima']} - {resumen['edad_maxima']} años"],
        ["Municipios con registros", str(resumen["total_municipios_con_registros"])],
    ]
    elementos.append(_tabla_estilizada(datos_resumen, anchos_col=[10 * cm, 6 * cm]))
    elementos.append(Spacer(1, 0.4 * cm))

    # --- Gráficos (si se solicitan) ---
    archivos_temporales = []
    if incluir_graficos:
        from controllers import graficos_controller as gc

        graficos_a_incluir = [
            ("Distribución por Género", gc.grafico_distribucion_genero, {}),
            ("Distribución por Edad", gc.grafico_distribucion_edad, {}),
            ("Distribución por Nivel Educativo", gc.grafico_nivel_educativo, {}),
            ("Participación por Municipio", gc.grafico_distribucion_municipio, {}),
            ("Distribución por Estatus", gc.grafico_distribucion_estatus, {}),
            ("Tendencia Temporal de Registros", gc.grafico_tendencia_temporal, {}),
        ]
        for titulo, funcion_grafico, kwargs in graficos_a_incluir:
            elementos.append(Paragraph(titulo, styles["SeccionTitulo"]))
            fig = funcion_grafico(**kwargs)
            nombre_temp = (
                titulo.replace(" ", "_").replace("é", "e").replace("ó", "o").replace("í", "i")
            )
            ruta_img_temp = os.path.join(config.REPORTS_DIR, f"_temp_{nombre_temp}.png")
            fig.savefig(ruta_img_temp, dpi=150, bbox_inches="tight")
            archivos_temporales.append(ruta_img_temp)
            elementos.append(Image(ruta_img_temp, width=14 * cm, height=8.5 * cm, kind="proportional"))
            elementos.append(Spacer(1, 0.3 * cm))

    # --- Distribución por municipio (tabla detallada) ---
    elementos.append(Paragraph("Detalle de Participación por Municipio", styles["SeccionTitulo"]))
    datos_municipio = [["Municipio", "Total Registros"]]
    for municipio, total in ec.distribucion_por_municipio().items():
        datos_municipio.append([municipio, str(total)])
    elementos.append(_tabla_estilizada(datos_municipio, anchos_col=[10 * cm, 6 * cm]))

    doc.build(elementos)

    # Limpieza de imágenes temporales generadas para el PDF
    for ruta_temp in archivos_temporales:
        try:
            os.remove(ruta_temp)
        except OSError:
            pass

    return ruta_salida


def generar_reporte_personal_pdf(
    lista_personal: List[PersonalMilitar],
    titulo_reporte: str = "Listado de Personal Militar",
    ruta_salida: Optional[str] = None,
) -> str:
    """
    Genera un PDF tabular con el listado de personal (cédula, nombre,
    municipio, estatus, etc.), paginado automáticamente por reportlab
    si la lista es extensa.
    """
    if ruta_salida is None:
        nombre_archivo = f"listado_personal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta_salida = os.path.join(config.REPORTS_DIR, nombre_archivo)

    styles = _estilos_pdf()
    doc = SimpleDocTemplate(
        ruta_salida, pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
    )
    elementos = _encabezado_institucional(styles)
    elementos.append(Paragraph(titulo_reporte, styles["SeccionTitulo"]))
    elementos.append(Paragraph(f"Total de registros: {len(lista_personal)}", styles["Normal"]))
    elementos.append(Spacer(1, 0.3 * cm))

    datos = [["Cédula", "Nombre Completo", "Edad", "Género", "Municipio", "Estatus"]]
    for p in lista_personal:
        datos.append([
            p.cedula, f"{p.nombres} {p.apellidos}", str(p.edad),
            p.genero, p.municipio, p.estatus,
        ])

    tabla = _tabla_estilizada(
        datos, anchos_col=[2.2 * cm, 5.5 * cm, 1.5 * cm, 2.3 * cm, 3 * cm, 3 * cm]
    )
    elementos.append(tabla)
    doc.build(elementos)
    return ruta_salida
